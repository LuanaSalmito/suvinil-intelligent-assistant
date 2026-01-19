"""
Script de inicialização do banco de dados.

Cria:
- Usuários de exemplo (admin e user)
- Catálogo de tintas Suvinil (carregado do arquivo XLSX)
"""
import sys
from pathlib import Path

# Adicionar root ao path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from sqlalchemy.orm import Session
from app.core.database import engine, Base, SessionLocal
from app.core.security import get_password_hash
from app.models.user import User, UserRole
from app.models.paint import Paint, Ambiente, Acabamento, Linha


def load_tintas_from_xlsx() -> list:
    """Carrega tintas do arquivo XLSX"""
    xlsx_path = Path(__file__).parent.parent / "docs" / "Base_de_Dados_de_Tintas_Suvinil.xlsx"
    
    if not xlsx_path.exists():
        print(f"      AVISO: Arquivo XLSX não encontrado em {xlsx_path}")
        return []
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        tintas = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = row
                continue
            
            # Mapear valores do XLSX para os enums
            ambiente_map = {
                "Interno": Ambiente.INTERNO,
                "Externo": Ambiente.EXTERNO,
                "Interno/Externo": Ambiente.INTERNO_EXTERNO,
            }
            
            acabamento_map = {
                "Fosco": Acabamento.FOSCO,
                "Acetinado": Acabamento.ACETINADO,
                "Brilhante": Acabamento.BRILHANTE,
            }
            
            linha_map = {
                "Premium": Linha.PREMIUM,
                "Standard": Linha.STANDARD,
            }
            
            # Extrair valores da linha
            nome = row[0]
            cor = row[1]
            tipo_parede = row[2]
            ambiente_str = row[3]
            acabamento_str = row[4]
            features = row[5]
            linha_str = row[6]
            
            if not nome:  # Pular linhas vazias
                continue
            
            tinta = {
                "nome": nome,
                "cor": cor,
                "tipo_parede": tipo_parede,
                "ambiente": ambiente_map.get(ambiente_str, Ambiente.INTERNO),
                "acabamento": acabamento_map.get(acabamento_str, Acabamento.FOSCO),
                "features": features,
                "linha": linha_map.get(linha_str, Linha.STANDARD),
            }
            tintas.append(tinta)
        
        return tintas
        
    except ImportError:
        print("      AVISO: openpyxl não instalado. Execute: pip install openpyxl")
        return []
    except Exception as e:
        print(f"      ERRO ao carregar XLSX: {e}")
        return []


def init_db():
    """Inicializa banco de dados com dados de exemplo"""
    print("=" * 60)
    print("   SUVINIL AI - Inicialização do Banco de Dados")
    print("=" * 60)
    
    print("\n[1/3] Criando tabelas...")
    Base.metadata.create_all(bind=engine)
    print("      Tabelas criadas com sucesso!")
    
    db = SessionLocal()
    try:
        # ========================================
        # CRIAR USUÁRIOS
        # ========================================
        admin = db.query(User).filter(User.username == "admin").first()
        
        if not admin:
            print("\n[2/3] Criando usuários de exemplo...")
            
            # Criar admin
            admin = User(
                email="admin@suvinil.com",
                username="admin",
                hashed_password=get_password_hash("admin123"),
                full_name="Administrador Suvinil",
                role=UserRole.ADMIN,
                is_active=True,
            )
            db.add(admin)
            
            # Criar usuário comum
            user = User(
                email="user@suvinil.com",
                username="user",
                hashed_password=get_password_hash("user123"),
                full_name="Usuário Teste",
                role=UserRole.USER,
                is_active=True,
            )
            db.add(user)
            
            # Criar usuário de demonstração
            demo = User(
                email="demo@suvinil.com",
                username="demo",
                hashed_password=get_password_hash("demo123"),
                full_name="Usuário Demonstração",
                role=UserRole.USER,
                is_active=True,
            )
            db.add(demo)
            
            db.commit()
            print("      Usuários criados:")
            print("      - Admin: admin / admin123")
            print("      - User:  user / user123")
            print("      - Demo:  demo / demo123")
        else:
            print("\n[2/3] Usuários já existem. Pulando...")
        
        # ========================================
        # CRIAR CATÁLOGO DE TINTAS (do XLSX)
        # ========================================
        paint_count = db.query(Paint).count()
        
        if paint_count == 0:
            tintas_catalogo = load_tintas_from_xlsx()
            
            if tintas_catalogo:
                print(f"\n[3/3] Criando catálogo de tintas ({len(tintas_catalogo)} produtos do XLSX)...")
                
                # Buscar usuário admin para created_by
                admin_user = db.query(User).filter(User.username == "admin").first()
                admin_id = admin_user.id if admin_user else None
                
                for i, paint_data in enumerate(tintas_catalogo, 1):
                    paint = Paint(
                        **paint_data,
                        is_active=True,
                        created_by=admin_id,
                    )
                    db.add(paint)
                
                db.commit()
                print(f"      Catálogo completo: {len(tintas_catalogo)} tintas criadas!")
                
                # Mostrar resumo do catálogo
                print("\n      Resumo do catálogo:")
                print(f"      - Linha Premium: {len([t for t in tintas_catalogo if t['linha'] == Linha.PREMIUM])} produtos")
                print(f"      - Linha Standard: {len([t for t in tintas_catalogo if t['linha'] == Linha.STANDARD])} produtos")
                print(f"      - Interno: {len([t for t in tintas_catalogo if t['ambiente'] == Ambiente.INTERNO])} produtos")
                print(f"      - Externo: {len([t for t in tintas_catalogo if t['ambiente'] == Ambiente.EXTERNO])} produtos")
                print(f"      - Interno/Externo: {len([t for t in tintas_catalogo if t['ambiente'] == Ambiente.INTERNO_EXTERNO])} produtos")
            else:
                print("\n[3/3] Não foi possível carregar tintas do XLSX. Pulando...")
        else:
            print(f"\n[3/3] Banco já possui {paint_count} tintas. Pulando...")
        
        print("\n" + "=" * 60)
        print("   Banco de dados inicializado com sucesso!")
        print("=" * 60)
        
    except Exception as e:
        db.rollback()
        print(f"\n   ERRO ao inicializar banco: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    init_db()
